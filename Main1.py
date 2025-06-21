from fastapi import FastAPI
from fastapi.responses import FileResponse, HTMLResponse
import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

app = FastAPI()

DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "Janan!1802",
    "database": "my_db"
}

@app.get("/", response_class=HTMLResponse)
def root():
    return """
    <html>
        <body style="font-family: Arial; padding: 40px;">
            <h2>Export Global RACT File</h2>
            <form action="/download" method="get">
                <button style="padding: 10px 20px; font-size: 16px;">Download Excel</button>
            </form>
        </body>
    </html>
    """

@app.get("/download")
def download_excel():
    # Load column mapping
    with open("Excel_Mapping.json") as f:
        column_map = json.load(f)

    # Connect to MySQL and fetch data
    conn = mysql.connector.connect(**DB_CONFIG)
    df = pd.read_sql("SELECT * FROM global_ract", conn)
    conn.close()

    # Apply column mapping to match Excel headers
    df.rename(columns=column_map, inplace=True)

    template_path = "RACT_Template.xlsx"
    export_path = "Instructions for BXU579481_Rev B.xlsx"
    sheet_name = "Appendix A GLOBAL RACT"

    # Load template
    workbook = load_workbook(template_path)
    worksheet = workbook[sheet_name]

    # Write DataFrame content starting from row 3
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, col in enumerate(df.columns, start=1):
        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len

    # Save as a new file (preserves original template)
    workbook.save(export_path)

    return FileResponse(
        path=export_path,
        filename="Instructions for BXU579481_Rev B.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
