from fastapi import APIRouter
from fastapi.responses import FileResponse
from models import engine  # Reuse your existing DB engine
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import os
from table_creation import content_creation
from table_conversion import conversion
from Update_table import update_appendix_b_from_csv

router = APIRouter()

@router.get("/ract/download")
def download_ract_excel():
    try:
        content_creation()
        db_config = {
            'dbname': 'baxterdb',
            'user': 'baxterdev',
            'password': 'Password123',
            'host': 'db',
            'port': '5432'
        }
        conversion()
        update_appendix_b_from_csv('Appendix_B.csv', db_config)

        # Load Excel column mappings for each table
        with open("./ractfile/Excel_Mapping.json") as f:
            column_map_a = json.load(f)
        with open("./ractfile/appendixB.json") as f:
            column_map_b = json.load(f)
        with open("./ractfile/appendixC.json") as f:
            column_map_c = json.load(f)

        # Table and mapping config
        table_sheet_map = [
            {
                "table": "appendix_a_global_ract",
                "sheet": "Appendix A GLOBAL RACT",
                "mapping": column_map_a,
                "start_row": 3  # Start from row 3, no header
            },
            {
                "table": "appendix_b_p2_conversion",
                "sheet": "Appendix B - P2 Conversion",
                "mapping": column_map_b,
                "start_row": 3  # Start from row 3
            },
            {
                "table": "Appendix_C_P1",
                "sheet": "Appendix C - P1 Table",
                "mapping": column_map_c,
                "start_row": 2  # Start from row 2
            }
        ]

        template_path = "./ractfile/RACT_Template.xlsx"
        export_path = "Instructions for BXU579481_Rev B.xlsx"
        workbook = load_workbook(template_path)

        for idx, config in enumerate(table_sheet_map):
            df = pd.read_sql(f'SELECT * FROM "{config["table"]}"', engine)
            df.rename(columns=config["mapping"], inplace=True)
            worksheet = workbook[config["sheet"]]

            # For first two tables: start from row 3, no header
            if idx in [0, 1]:
                for row_idx, row in enumerate(df.itertuples(index=False), start=3):
                    for col_idx, value in enumerate(row, start=1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
            # For third table: start from row 2, no header
            elif idx == 2:
                for row_idx, row in enumerate(df.itertuples(index=False), start=2):
                    for col_idx, value in enumerate(row, start=1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col_idx, col in enumerate(df.columns, start=1):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len

        # Save to file
        workbook.save(export_path)

        return FileResponse(
            path=export_path,
            filename=os.path.basename(export_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return {"error": str(e)}
